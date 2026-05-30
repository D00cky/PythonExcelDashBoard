from collections.abc import Iterable
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
from openpyxl.workbook import Workbook

DATA_SHEET_PREFIX = "DADOS - "
COVER_SHEET = "CAPA"
PERIODO_CELL = "B4"
PERIODO_PREFIX = "Período: "

IC_SERVICE_ROWS = {
    "Água": 11,
    "Esgoto": 12,
    "Reposição": 13,
}

IQS_SERVICE_ROWS = {
    "Água": 26,
    "Esgoto": 27,
    "Cavalete": 28,
    "Reposição": 29,
}
IQS_TOTAL_ROW = 30
IQS_OVERALL_CELL = f"G{IQS_TOTAL_ROW}"


@dataclass(frozen=True)
class ServiceIC:
    name: str
    ic_pct: float
    lvs: int


@dataclass(frozen=True)
class ServiceIQS:
    name: str
    fotos_avaliadas: int
    fotos_nc: int
    fotos_conforme: int
    nc_pct: float
    conforme_pct: float


class SabespPimentasTemplate:
    SERVICE_SHEETS = frozenset({"ÁGUA", "ESGOTO", "CAVALETE", "REPOSIÇÃO"})
    MIN_SERVICES = 2

    def __init__(self, data_sheet_name: str = "DADOS - PIMENTAS"):
        self.data_sheet_name = data_sheet_name
        self.polo_name = data_sheet_name.removeprefix(DATA_SHEET_PREFIX).strip()

    @classmethod
    def detect(cls, sheet_names: Iterable[str]) -> "SabespPimentasTemplate | None":
        names = {name.strip() for name in sheet_names}
        if COVER_SHEET not in names:
            return None
        data_sheet = next((n for n in names if n.startswith(DATA_SHEET_PREFIX)), None)
        if data_sheet is None:
            return None
        if len(cls.SERVICE_SHEETS & names) < cls.MIN_SERVICES:
            return None
        return cls(data_sheet_name=data_sheet)

    @classmethod
    def matches(cls, sheet_names: Iterable[str]) -> bool:
        return cls.detect(sheet_names) is not None

    def extract_periodo(self, workbook: Workbook) -> str | None:
        if self.data_sheet_name not in workbook.sheetnames:
            return None
        value = workbook[self.data_sheet_name][PERIODO_CELL].value
        if not isinstance(value, str):
            return None
        return value.removeprefix(PERIODO_PREFIX).strip() or None

    def extract_iqs_by_service(self, workbook: Workbook) -> list[ServiceIQS]:
        if self.data_sheet_name not in workbook.sheetnames:
            return []
        ws = workbook[self.data_sheet_name]
        rows: list[ServiceIQS] = []
        for name, row in IQS_SERVICE_ROWS.items():
            if ws[f"B{row}"].value != name:
                continue
            rows.append(
                ServiceIQS(
                    name=name,
                    fotos_avaliadas=ws[f"C{row}"].value or 0,
                    fotos_nc=ws[f"D{row}"].value or 0,
                    fotos_conforme=ws[f"E{row}"].value or 0,
                    nc_pct=ws[f"F{row}"].value or 0.0,
                    conforme_pct=ws[f"G{row}"].value or 0.0,
                )
            )
        return rows

    def extract_iqs_overall(self, workbook: Workbook) -> float | None:
        if self.data_sheet_name not in workbook.sheetnames:
            return None
        value = workbook[self.data_sheet_name][IQS_OVERALL_CELL].value
        return value if isinstance(value, int | float) else None

    def extract_ic_by_service(self, workbook: Workbook) -> list[ServiceIC]:
        if self.data_sheet_name not in workbook.sheetnames:
            return []
        ws = workbook[self.data_sheet_name]
        rows: list[ServiceIC] = []
        for name, row in IC_SERVICE_ROWS.items():
            if ws[f"B{row}"].value != name:
                continue
            rows.append(
                ServiceIC(
                    name=name,
                    ic_pct=ws[f"C{row}"].value or 0.0,
                    lvs=ws[f"E{row}"].value or 0,
                )
            )
        return rows

    def extract_team_detail(self, path: Path, team_name: str) -> dict[str, dict]:
        """Per-service breakdown for one team, at the OS (inspection) level.

        Returns {service: {inspecoes, conforme, nao_conforme, with_nc,
        with_sf, top_nc_reason, top_sf_reason, tss_summary,
        failing_inspections}}.

        ``conforme + nao_conforme == inspecoes``. An inspection is conforme
        iff every stage cell is 'C' (NA / blank do not count as failures).
        ``with_nc`` / ``with_sf`` are informational subsets of
        ``nao_conforme``: inspections with at least one NC / SF stage.
        ``failing_inspections`` lists every not-conforme OS with TSS and
        the stages that failed plus their observation text when present.
        """
        target = team_name.strip()
        result: dict[str, dict] = {}
        sheets = self._service_sheets(path)
        for service in sorted(self.SERVICE_SHEETS):
            df = sheets.get(service)
            if df is None:
                continue
            team_col = _ci_column(df, "EQUIPE")
            if team_col is None:
                continue
            tss_col = _ci_column(df, "Descrição TSS")
            sub = df[df[team_col].astype(str).str.strip() == target]
            if sub.empty:
                continue

            pairs = _stage_observation_pairs(sub, exclude={team_col, tss_col})
            os_col = _ci_column(df, "Número OS")
            result[service] = {
                "inspecoes": int(len(sub)),
                **_os_conformity_summary(sub, [s for s, _ in pairs]),
                "top_nc_reason": _top_reason(_collect_reasons(sub, pairs, "NC")),
                "top_sf_reason": _top_reason(_collect_reasons(sub, pairs, "SF")),
                "tss_summary": _top_tss(sub, tss_col),
                "failing_inspections": _failing_inspections(sub, pairs, tss_col, os_col),
            }
        return result

    def extract_inspections(self, path: Path) -> pd.DataFrame:
        """Read the four service sheets, return one row per inspection.

        Columns: team, tss, service, conforme_count, nao_conforme_count,
        start_date. ``conforme_count`` / ``nao_conforme_count`` are 0/1
        per-row indicators at the OS (inspection) level: an inspection is
        conforme iff every stage cell is ``C`` (``NA`` and blank are
        treated as "not a failure"). Any ``NC`` or ``SF`` makes the whole
        row not-conforme. Summing these columns therefore yields OS
        totals that always equal the number of inspections — not cell
        counts. start_date comes from 'Data Início Execução' and is
        coerced to datetime.
        """
        return _inspections_cached(
            str(path), path.stat().st_mtime_ns, tuple(sorted(self.SERVICE_SHEETS))
        )

    def _service_sheets(self, path: Path) -> dict[str, pd.DataFrame]:
        return _read_service_sheets_cached(
            str(path), path.stat().st_mtime_ns, tuple(sorted(self.SERVICE_SHEETS))
        )

    def build_service_iqs_bar(self, rows: list[ServiceIQS]) -> go.Figure:
        return go.Figure(
            data=[
                go.Bar(
                    x=[r.name for r in rows],
                    y=[r.conforme_pct for r in rows],
                    marker_color="#2a9d8f",
                    text=[f"{r.conforme_pct:.1%}" for r in rows],
                    textposition="outside",
                )
            ],
            layout=go.Layout(
                title="Índice de Qualidade SABESP por Serviço",
                yaxis={"title": "Conforme (%)", "tickformat": ".0%", "range": [0, 1]},
                xaxis={"title": "Serviço"},
                template="plotly_white",
            ),
        )

    def build_ic_bar(self, rows: list[ServiceIC]) -> go.Figure:
        return go.Figure(
            data=[
                go.Bar(
                    x=[r.name for r in rows],
                    y=[r.ic_pct for r in rows],
                    marker_color="#264653",
                    text=[f"{r.ic_pct:.0%}" for r in rows],
                    textposition="outside",
                )
            ],
            layout=go.Layout(
                title="Índice de Conformidade (IC) por Serviço",
                yaxis={"title": "IC (%)", "tickformat": ".0%", "range": [0, 1.1]},
                xaxis={"title": "Serviço"},
                template="plotly_white",
            ),
        )

    def build_photo_conformity_stacked(self, rows: list[ServiceIQS]) -> go.Figure:
        names = [r.name for r in rows]
        return go.Figure(
            data=[
                go.Bar(
                    name="Não Conforme",
                    x=names,
                    y=[r.fotos_nc for r in rows],
                    marker_color="#e76f51",
                ),
                go.Bar(
                    name="Conforme",
                    x=names,
                    y=[r.fotos_conforme for r in rows],
                    marker_color="#2a9d8f",
                ),
            ],
            layout=go.Layout(
                barmode="stack",
                title="Fotos Avaliadas por Serviço",
                yaxis={"title": "Quantidade de fotos"},
                xaxis={"title": "Serviço"},
                template="plotly_white",
            ),
        )

    def build_tss_distribution(self, df: pd.DataFrame, top_n: int = 15) -> go.Figure:
        if df.empty:
            return _empty_figure("Sem inspeções registradas")

        counts = df.groupby("tss").size().sort_values(ascending=False).head(top_n)
        labels = list(reversed(counts.index))
        values = list(reversed(counts.values))

        return go.Figure(
            data=[
                go.Bar(
                    x=values,
                    y=labels,
                    orientation="h",
                    marker_color="#457b9d",
                    text=values,
                    textposition="outside",
                )
            ],
            layout=go.Layout(
                title=f"Tipos de Serviço (TSS) — Top {len(labels)}",
                xaxis={"title": "Quantidade de inspeções"},
                yaxis={"title": "Descrição TSS", "automargin": True},
                template="plotly_white",
                height=max(400, 30 * len(labels) + 100),
                margin={"l": 280},
            ),
        )

    def build_team_conformity_for_service(
        self, df: pd.DataFrame, service: str, top_n: int = 10
    ) -> go.Figure:
        return _conformity_chart(df, service, "team", top_n, "Equipe")

    def build_tss_conformity_for_service(
        self, df: pd.DataFrame, service: str, top_n: int = 10
    ) -> go.Figure:
        return _conformity_chart(df, service, "tss", top_n, "TSS")

    def build_team_service_stacked(self, df: pd.DataFrame, top_n: int = 15) -> go.Figure:
        if df.empty:
            return _empty_figure("Sem inspeções registradas")

        team_totals = df.groupby("team").size().sort_values(ascending=False).head(top_n)
        teams = list(reversed(team_totals.index))  # plotly y starts at bottom

        pivot = (
            df[df["team"].isin(teams)]
            .groupby(["team", "service"])
            .size()
            .unstack(fill_value=0)
            .reindex(teams)
        )

        traces = []
        for service in sorted(self.SERVICE_SHEETS):
            if service not in pivot.columns:
                continue
            traces.append(
                go.Bar(
                    name=service,
                    y=teams,
                    x=pivot[service].tolist(),
                    orientation="h",
                    marker_color=_SERVICE_COLORS.get(service, "#888"),
                )
            )

        return go.Figure(
            data=traces,
            layout=go.Layout(
                barmode="stack",
                title=f"Inspeções por Equipe (Top {len(teams)})",
                xaxis={"title": "Quantidade de inspeções"},
                yaxis={"title": "Equipe", "automargin": True},
                template="plotly_white",
                height=max(400, 30 * len(teams) + 100),
                legend={"orientation": "h", "yanchor": "bottom", "y": 1.02},
            ),
        )


@lru_cache(maxsize=8)
def _read_service_sheets_cached(
    path_str: str, mtime_ns: int, services: tuple[str, ...]
) -> dict[str, pd.DataFrame]:
    xf = pd.ExcelFile(path_str, engine="calamine")
    out: dict[str, pd.DataFrame] = {}
    for service in services:
        if service not in xf.sheet_names:
            continue
        df = xf.parse(service)
        df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
        out[service] = df
    return out


_INSPECTIONS_COLUMNS = (
    "team",
    "tss",
    "service",
    "conforme_count",
    "nao_conforme_count",
    "photo_conforme",
    "photo_nc",
    "photo_sf",
    "photo_total",
    "start_date",
)


@lru_cache(maxsize=8)
def _inspections_cached(
    path_str: str, mtime_ns: int, services: tuple[str, ...]
) -> pd.DataFrame:
    sheets = _read_service_sheets_cached(path_str, mtime_ns, services)
    parts: list[pd.DataFrame] = []
    for service in services:
        df = sheets.get(service)
        if df is None:
            continue
        team_col = _ci_column(df, "EQUIPE")
        tss_col = _ci_column(df, "Descrição TSS")
        if team_col is None or tss_col is None:
            continue
        df = df.dropna(subset=[team_col, tss_col]).copy()
        stage_cols = _detect_stage_columns(df, exclude={team_col, tss_col})
        if stage_cols:
            stages = df[stage_cols].astype(str).apply(lambda s: s.str.strip())
            is_failing = stages.isin(_FAILING_STAGE_CODES).any(axis=1)
            df["conforme_count"] = (~is_failing).astype(int)
            df["nao_conforme_count"] = is_failing.astype(int)
            df["photo_conforme"] = (stages == "C").sum(axis=1).astype(int)
            df["photo_nc"] = (stages == "NC").sum(axis=1).astype(int)
            df["photo_sf"] = (stages == "SF").sum(axis=1).astype(int)
            df["photo_total"] = (
                df["photo_conforme"] + df["photo_nc"] + df["photo_sf"]
            )
        else:
            df["conforme_count"] = 0
            df["nao_conforme_count"] = 0
            df["photo_conforme"] = 0
            df["photo_nc"] = 0
            df["photo_sf"] = 0
            df["photo_total"] = 0
        date_col = _ci_column(df, "Data Início Execução")
        df["start_date"] = (
            pd.to_datetime(df[date_col], errors="coerce") if date_col is not None else pd.NaT
        )
        df = df.rename(columns={team_col: "team", tss_col: "tss"})
        df = df[
            [
                "team",
                "tss",
                "conforme_count",
                "nao_conforme_count",
                "photo_conforme",
                "photo_nc",
                "photo_sf",
                "photo_total",
                "start_date",
            ]
        ].copy()
        df["service"] = service
        parts.append(df)
    if not parts:
        return pd.DataFrame(columns=list(_INSPECTIONS_COLUMNS))
    return pd.concat(parts, ignore_index=True)


_SERVICE_COLORS = {
    "ÁGUA": "#2a9d8f",
    "ESGOTO": "#264653",
    "CAVALETE": "#e9c46a",
    "REPOSIÇÃO": "#e76f51",
}


def _empty_figure(message: str) -> go.Figure:
    return go.Figure(
        layout=go.Layout(
            template="plotly_white",
            annotations=[
                {
                    "text": message,
                    "showarrow": False,
                    "xref": "paper",
                    "yref": "paper",
                    "x": 0.5,
                    "y": 0.5,
                    "font": {"size": 14, "color": "#888"},
                }
            ],
        )
    )


def _ci_column(df: pd.DataFrame, target: str) -> str | None:
    target_lower = target.casefold()
    for col in df.columns:
        if isinstance(col, str) and col.casefold() == target_lower:
            return col
    return None


_STAGE_CODE_VALUES = frozenset({"C", "NC", "SF", "NA"})
_FAILING_STAGE_CODES = frozenset({"NC", "SF"})


def _conformity_chart(
    df: pd.DataFrame, service: str, dim_col: str, top_n: int, dim_label: str
) -> go.Figure:
    sub = df[df["service"] == service] if not df.empty else df
    if sub.empty:
        return _empty_figure(f"Sem inspeções em {service}")

    grouped = sub.groupby(dim_col).agg(
        conforme=("conforme_count", "sum"),
        nao_conforme=("nao_conforme_count", "sum"),
    )
    grouped["total"] = grouped["conforme"] + grouped["nao_conforme"]
    grouped = grouped[grouped["total"] > 0].sort_values("total", ascending=False).head(top_n)
    if grouped.empty:
        return _empty_figure(f"Sem dados de conformidade em {service}")

    labels = list(reversed(grouped.index))
    conforme_vals = grouped.loc[labels, "conforme"].tolist()
    nc_vals = grouped.loc[labels, "nao_conforme"].tolist()

    return go.Figure(
        data=[
            go.Bar(
                name="Conforme",
                y=labels,
                x=conforme_vals,
                orientation="h",
                marker_color="#2a9d8f",
            ),
            go.Bar(
                name="Não Conforme",
                y=labels,
                x=nc_vals,
                orientation="h",
                marker_color="#e76f51",
            ),
        ],
        layout=go.Layout(
            barmode="stack",
            title=f"{service} — Conformidade por {dim_label}",
            xaxis={"title": "Etapas avaliadas"},
            yaxis={"title": dim_label, "automargin": True},
            template="plotly_white",
            height=max(350, 30 * len(labels) + 120),
            legend={"orientation": "h", "yanchor": "bottom", "y": 1.02},
        ),
    )


def _os_conformity_summary(sub: pd.DataFrame, stage_cols: list[str]) -> dict[str, int]:
    """OS-level conformity for a team-service slice.

    An OS is conforme iff none of its stages is 'NC' or 'SF'. ``with_nc``
    / ``with_sf`` count OSes that have at least one stage cell in that
    code — informational subsets of ``nao_conforme``.
    """
    if not stage_cols:
        return {"conforme": 0, "nao_conforme": 0, "with_nc": 0, "with_sf": 0}
    stages = sub[stage_cols].astype(str).apply(lambda s: s.str.strip())
    is_failing = stages.isin(_FAILING_STAGE_CODES).any(axis=1)
    return {
        "conforme": int((~is_failing).sum()),
        "nao_conforme": int(is_failing.sum()),
        "with_nc": int((stages == "NC").any(axis=1).sum()),
        "with_sf": int((stages == "SF").any(axis=1).sum()),
    }


def _failing_inspections(
    sub: pd.DataFrame,
    pairs: list[tuple[str, str | None]],
    tss_col: str | None,
    os_col: str | None = None,
) -> list[dict]:
    """Rows where at least one stage is NC or SF, with per-stage detail."""
    if not pairs:
        return []
    out: list[dict] = []
    for _, row in sub.iterrows():
        failed: list[dict] = []
        for stage_col, obs_col in pairs:
            val = str(row[stage_col]).strip() if pd.notna(row[stage_col]) else ""
            if val not in _FAILING_STAGE_CODES:
                continue
            observation = (
                str(row[obs_col]).strip()
                if obs_col is not None and pd.notna(row[obs_col])
                else None
            )
            failed.append({"stage": stage_col, "code": val, "observation": observation})
        if not failed:
            continue
        tss = (
            str(row[tss_col]).strip()
            if tss_col is not None and pd.notna(row[tss_col])
            else ""
        )
        out.append({"os": _format_os(row, os_col), "tss": tss, "failed_stages": failed})
    return out


def _format_os(row: pd.Series, os_col: str | None) -> str:
    if os_col is None or pd.isna(row[os_col]):
        return ""
    raw = row[os_col]
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw).strip()


def _stage_code_counts(sub: pd.DataFrame, stage_cols: list[str]) -> dict[str, int]:
    """Total occurrences of each stage code (C/NC/SF/NA) across stage columns."""
    if not stage_cols:
        return {}
    flat = sub[stage_cols].astype(str).apply(lambda s: s.str.strip()).values.ravel()
    return {str(k): int(v) for k, v in pd.Series(flat).value_counts().items()}


def _collect_reasons(
    sub: pd.DataFrame, pairs: list[tuple[str, str | None]], code: str
) -> list[str]:
    """Observation strings paired with rows where the stage cell equals ``code``."""
    out: list[str] = []
    for stage_col, obs_col in pairs:
        if obs_col is None:
            continue
        stage_norm = sub[stage_col].astype(str).str.strip()
        out.extend(sub.loc[stage_norm == code, obs_col].dropna().astype(str).tolist())
    return out


def _top_reason(reasons: list[str]) -> tuple[str, int] | None:
    if not reasons:
        return None
    top = pd.Series(reasons).value_counts().head(1)
    return (str(top.index[0]), int(top.iloc[0]))


def _top_tss(sub: pd.DataFrame, tss_col: str | None, n: int = 10) -> list[tuple[str, int]]:
    if tss_col is None:
        return []
    counts = sub[tss_col].dropna().astype(str).value_counts().head(n)
    return [(str(k), int(v)) for k, v in counts.items()]


def _stage_observation_pairs(df: pd.DataFrame, exclude: set) -> list[tuple[str, str | None]]:
    """For each detected stage column, pair it with the following column iff
    that column's name starts with 'Observa' (case-insensitive)."""
    stages = _detect_stage_columns(df, exclude=exclude)
    cols = list(df.columns)
    pairs: list[tuple[str, str | None]] = []
    for stage in stages:
        idx = cols.index(stage)
        candidate = cols[idx + 1] if idx + 1 < len(cols) else None
        if isinstance(candidate, str) and candidate.casefold().startswith("observa"):
            pairs.append((stage, candidate))
        else:
            pairs.append((stage, None))
    return pairs


def _detect_stage_columns(df: pd.DataFrame, exclude: set) -> list[str]:
    """Return columns whose every non-null value is one of {C, NC, SF, NA}."""
    cols: list[str] = []
    for col in df.columns:
        if col in exclude:
            continue
        non_null = df[col].dropna()
        if len(non_null) == 0:
            continue
        as_str = non_null.astype(str).str.strip()
        if as_str.isin(_STAGE_CODE_VALUES).all():
            cols.append(col)
    return cols
