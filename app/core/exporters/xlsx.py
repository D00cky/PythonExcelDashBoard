from io import BytesIO
from pathlib import Path

from openpyxl import Workbook as NewWorkbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook import Workbook

from app.core.templates.sabesp_pimentas import SabespPimentasTemplate

_HEADER_FILL = PatternFill("solid", fgColor="264653")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def render_xlsx(template: SabespPimentasTemplate, workbook: Workbook, path: Path) -> bytes:
    inspections = template.extract_inspections(path)
    iqs_rows = template.extract_iqs_by_service(workbook)
    ic_rows = template.extract_ic_by_service(workbook)
    iqs_overall = template.extract_iqs_overall(workbook)
    periodo = _periodo(inspections) or template.extract_periodo(workbook)

    wb = NewWorkbook()
    _write_summary(
        wb.active,
        polo=template.polo_name.title(),
        periodo=periodo,
        iqs_overall=iqs_overall,
        total_inspections=len(inspections),
        unique_teams=inspections["team"].nunique() if not inspections.empty else 0,
    )

    _write_table(
        wb.create_sheet("IC por Serviço"),
        headers=["Serviço", "IC (%)", "Quantidade de LVs"],
        rows=[[r.name, r.ic_pct, r.lvs] for r in ic_rows],
        pct_cols={1},
    )
    _write_table(
        wb.create_sheet("IQS por Serviço"),
        headers=[
            "Serviço", "Fotos Avaliadas", "NC", "Conforme", "NC (%)", "Conforme (%)",
        ],
        rows=[
            [r.name, r.fotos_avaliadas, r.fotos_nc, r.fotos_conforme, r.nc_pct, r.conforme_pct]
            for r in iqs_rows
        ],
        pct_cols={4, 5},
    )

    if not inspections.empty:
        for service in sorted(template.SERVICE_SHEETS):
            sub = inspections[inspections["service"] == service]
            if sub.empty:
                continue
            agg = sub.groupby("team").agg(
                conforme=("conforme_count", "sum"),
                nao_conforme=("nao_conforme_count", "sum"),
            )
            agg["total"] = agg["conforme"] + agg["nao_conforme"]
            top = agg[agg["total"] > 0].sort_values("total", ascending=False).head(50)
            _write_table(
                wb.create_sheet(f"Equipes — {service}"[:31]),
                headers=["Equipe", "Conforme", "Não Conforme", "Total"],
                rows=[
                    [team, int(r["conforme"]), int(r["nao_conforme"]), int(r["total"])]
                    for team, r in top.iterrows()
                ],
            )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary(ws, *, polo: str, periodo, iqs_overall, total_inspections, unique_teams):
    ws.title = "Resumo"
    ws["A1"] = f"Dashboard SABESP — Polo {polo}"
    ws["A1"].font = Font(size=16, bold=True, color="264653")
    rows = [
        ("Período", periodo or "—"),
        ("IQS Geral", f"{iqs_overall:.1%}" if iqs_overall is not None else "—"),
        ("Total de inspeções", total_inspections),
        ("Equipes distintas", unique_teams),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28


def _write_table(ws, *, headers, rows, pct_cols: set[int] | None = None):
    pct_cols = pct_cols or set()
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row):
            cell = ws.cell(row=r, column=c + 1, value=value)
            if c in pct_cols and isinstance(value, int | float):
                cell.number_format = "0.0%"
    for c in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = max(
            18, len(str(headers[c - 1])) + 4
        )


def _periodo(df) -> str | None:
    if df.empty or "start_date" not in df.columns:
        return None
    dates = df["start_date"].dropna()
    if dates.empty:
        return None
    return f"{dates.min():%d/%m/%Y} à {dates.max():%d/%m/%Y}"
