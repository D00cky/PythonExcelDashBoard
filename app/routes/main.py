import uuid
from pathlib import Path
from typing import Any, Literal

import pandas as pd
from flask import (
    Blueprint,
    Response,
    abort,
    current_app,
    redirect,
    render_template,
    request,
    url_for,
)
from openpyxl import load_workbook

from app.core.exporters import render_export
from app.core.templates import recognize
from app.core.templates.sabesp_pimentas import (
    SabespPimentasTemplate,
    ServiceIC,
    ServiceIQS,
)

bp = Blueprint("main", __name__)


@bp.get("/")
def index() -> str:
    return render_template("index.html")


@bp.post("/upload")
def upload():
    if "file" not in request.files:
        abort(400)
    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        abort(400)

    upload_id = uuid.uuid4().hex
    uploads_dir = Path(current_app.instance_path) / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    file.save(uploads_dir / f"{upload_id}.xlsx")

    return redirect(url_for("main.dashboard", upload_id=upload_id), code=303)


@bp.get("/dashboard/<upload_id>")
def dashboard(upload_id: str) -> str:
    path = _upload_path(upload_id)
    workbook = load_workbook(path, data_only=True, read_only=True)
    template = recognize(workbook.sheetnames)
    if not isinstance(template, SabespPimentasTemplate):
        return render_template("dashboard_unknown.html", sheet_names=workbook.sheetnames)

    filter_start = _parse_iso_date(request.args.get("start", ""))
    filter_end = _parse_iso_date(request.args.get("end", ""))
    swap_dates = request.args.get("swap") == "1"

    return render_template(
        "dashboard.html",
        download_action=url_for("main.download", upload_id=upload_id),
        **_build_sabesp_context(
            template,
            workbook,
            path,
            plotly_mode="cdn",
            filter_start=filter_start,
            filter_end=filter_end,
            swap_dates=swap_dates,
        ),
    )


@bp.get("/dashboard/<upload_id>/team")
def team_detail(upload_id: str) -> str:
    team_name = (request.args.get("name") or "").strip()
    if not team_name:
        abort(400)
    path = _upload_path(upload_id)
    workbook = load_workbook(path, data_only=True, read_only=True)
    template = recognize(workbook.sheetnames)
    if not isinstance(template, SabespPimentasTemplate):
        abort(404)
    detail = template.extract_team_detail(path, team_name)
    if not detail:
        abort(404)
    return render_template(
        "team_detail.html",
        polo_name=template.polo_name.title(),
        team_name=team_name,
        dashboard_url=url_for("main.dashboard", upload_id=upload_id),
        detail=detail,
    )


_SUPPORTED_FORMATS = {"md", "xlsx", "pdf", "docx"}


@bp.get("/download/<upload_id>")
def download(upload_id: str) -> Response:
    fmt = request.args.get("fmt", "md").lower()
    if fmt not in _SUPPORTED_FORMATS:
        abort(400)

    path = _upload_path(upload_id)
    workbook = load_workbook(path, data_only=True, read_only=True)
    template = recognize(workbook.sheetnames)
    if not isinstance(template, SabespPimentasTemplate):
        abort(404)

    body, mimetype = render_export(fmt, template, workbook, path)
    response = Response(body, mimetype=mimetype)
    response.headers["Content-Disposition"] = f'attachment; filename="dashboard-{upload_id}.{fmt}"'
    return response


def _upload_path(upload_id: str) -> Path:
    path = Path(current_app.instance_path) / "uploads" / f"{upload_id}.xlsx"
    if not path.exists():
        abort(404)
    return path


def _periodo_from_inspections(df) -> str | None:
    if df.empty or "start_date" not in df.columns:
        return None
    dates = df["start_date"].dropna()
    if dates.empty:
        return None
    return f"{dates.min():%d/%m/%Y} à {dates.max():%d/%m/%Y}"


def _parse_iso_date(value: str) -> pd.Timestamp | None:
    """Parse YYYY-MM-DD from an <input type=date>; return None when invalid."""
    if not value:
        return None
    try:
        return pd.Timestamp(value)
    except (ValueError, TypeError):
        return None


_SUSPICIOUS_SPAN_DAYS = 60


def _build_sabesp_context(
    template: SabespPimentasTemplate,
    workbook,
    path: Path,
    plotly_mode: Literal["cdn", "inline"],
    filter_start: pd.Timestamp | None = None,
    filter_end: pd.Timestamp | None = None,
    swap_dates: bool = False,
) -> dict[str, Any]:
    full_inspections = template.extract_inspections(path)
    if swap_dates:
        full_inspections = _swap_day_month(full_inspections)

    available_start, available_end = _date_bounds(full_inspections)
    span_warning = _date_span_warning(available_start, available_end, swapped=swap_dates)
    inspections = _apply_date_filter(full_inspections, filter_start, filter_end)
    is_filtered = filter_start is not None or filter_end is not None
    recomputed = is_filtered or swap_dates

    if recomputed:
        services = sorted(template.SERVICE_SHEETS)
        iqs_rows = _iqs_rows_from_inspections(inspections, services)
        ic_rows = _ic_rows_from_inspections(inspections, services)
        iqs_overall = _iqs_overall_from_inspections(inspections)
    else:
        iqs_rows = template.extract_iqs_by_service(workbook)
        ic_rows = template.extract_ic_by_service(workbook)
        iqs_overall = template.extract_iqs_overall(workbook)

    periodo = _periodo_from_inspections(inspections) or template.extract_periodo(workbook)
    teams_sorted = (
        sorted(inspections["team"].dropna().unique().tolist()) if not inspections.empty else []
    )

    per_service_sections = []
    for idx, service in enumerate(sorted(template.SERVICE_SHEETS)):
        per_service_sections.append(
            {
                "service": service,
                "team_chart": template.build_team_conformity_for_service(
                    inspections, service
                ).to_html(include_plotlyjs=False, full_html=False, div_id=f"conf-team-{idx}"),
                "tss_chart": template.build_tss_conformity_for_service(
                    inspections, service
                ).to_html(include_plotlyjs=False, full_html=False, div_id=f"conf-tss-{idx}"),
            }
        )

    return {
        "polo_name": template.polo_name.title(),
        "periodo": periodo,
        "iqs_overall": iqs_overall,
        "total_fotos": sum(r.fotos_avaliadas for r in iqs_rows),
        "total_inspections": len(inspections),
        "fig_ic_bar": template.build_ic_bar(ic_rows).to_html(
            include_plotlyjs=plotly_mode, full_html=False, div_id="ic-bar"
        ),
        "fig_iqs_bar": template.build_service_iqs_bar(iqs_rows).to_html(
            include_plotlyjs=False, full_html=False, div_id="iqs-bar"
        ),
        "fig_photos": template.build_photo_conformity_stacked(iqs_rows).to_html(
            include_plotlyjs=False, full_html=False, div_id="photos"
        ),
        "fig_team_service": template.build_team_service_stacked(inspections).to_html(
            include_plotlyjs=False, full_html=False, div_id="team-service"
        ),
        "fig_tss": template.build_tss_distribution(inspections).to_html(
            include_plotlyjs=False, full_html=False, div_id="tss-distribution"
        ),
        "per_service_sections": per_service_sections,
        "teams_sorted": teams_sorted,
        "filter_start": _iso(filter_start),
        "filter_end": _iso(filter_end),
        "available_start": _iso(available_start),
        "available_end": _iso(available_end),
        "is_filtered": is_filtered,
        "swap_dates": swap_dates,
        "recomputed": recomputed,
        "span_warning": span_warning,
    }


def _date_bounds(df: pd.DataFrame) -> tuple[pd.Timestamp | None, pd.Timestamp | None]:
    if df.empty or "start_date" not in df.columns:
        return None, None
    dates = df["start_date"].dropna()
    if dates.empty:
        return None, None
    return dates.min().normalize(), dates.max().normalize()


def _apply_date_filter(
    df: pd.DataFrame,
    start: pd.Timestamp | None,
    end: pd.Timestamp | None,
) -> pd.DataFrame:
    if df.empty or "start_date" not in df.columns:
        return df
    mask = pd.Series(True, index=df.index)
    if start is not None:
        mask &= df["start_date"] >= start
    if end is not None:
        # End is inclusive on the day — bump to end of day.
        mask &= df["start_date"] < (end + pd.Timedelta(days=1))
    return df[mask]


def _date_span_warning(
    start: pd.Timestamp | None, end: pd.Timestamp | None, swapped: bool = False
) -> str | None:
    if start is None or end is None:
        return None
    span_days = (end - start).days
    if span_days <= _SUSPICIOUS_SPAN_DAYS:
        return None
    if swapped:
        return (
            f"Datas com dia/mês invertidos: {span_days} dias "
            f"({start:%d/%m/%Y} → {end:%d/%m/%Y}). O resultado ainda parece "
            "incorreto; verifique a planilha original."
        )
    return (
        f"Atenção: as datas das inspeções abrangem {span_days} dias "
        f"({start:%d/%m/%Y} → {end:%d/%m/%Y}). Isso pode indicar dia/mês "
        "trocados na planilha original."
    )


def _swap_day_month(df: pd.DataFrame) -> pd.DataFrame:
    """Flip day/month only when the swap moves a row into the target month.

    SABESP reports occasionally store dates as MM/DD instead of DD/MM at
    data entry. Unambiguous rows (day > 12) reveal the file's actual
    target month; ambiguous rows (day ≤ 12) are swapped only when the
    swap puts them into that target month. This avoids breaking the
    correctly-entered dates that happen to have day ≤ 12.
    """
    if df.empty or "start_date" not in df.columns:
        return df
    dates = df["start_date"]
    unambiguous = dates.notna() & (dates.dt.day > 12)
    if not unambiguous.any():
        return df  # nothing tells us which month is "right"
    target_month = int(dates[unambiguous].dt.month.mode().iloc[0])

    df = df.copy()
    ambiguous = dates.notna() & (dates.dt.day <= 12)
    # Swap only when swap-day-to-month would yield the target month.
    swap_mask = ambiguous & (dates.dt.day == target_month) & (dates.dt.month != target_month)
    if not swap_mask.any():
        return df
    sub = dates[swap_mask]
    df.loc[swap_mask, "start_date"] = pd.to_datetime(
        {
            "year": sub.dt.year,
            "month": sub.dt.day,
            "day": sub.dt.month,
            "hour": sub.dt.hour,
            "minute": sub.dt.minute,
        }
    ).values
    return df


def _iqs_rows_from_inspections(df: pd.DataFrame, services: list[str]) -> list[ServiceIQS]:
    """Reconstruct ServiceIQS records from raw inspection cells.

    Photos are summed across stage cells per service. ``fotos_nc`` lumps
    NC + SF together to match how the CAPA-aggregated row treats failures.
    """
    out: list[ServiceIQS] = []
    if df.empty:
        return out
    for svc in services:
        sub = df[df["service"] == svc]
        if sub.empty:
            continue
        avaliadas = int(sub["photo_total"].sum())
        if avaliadas == 0:
            continue
        nc = int((sub["photo_nc"] + sub["photo_sf"]).sum())
        conforme = int(sub["photo_conforme"].sum())
        out.append(
            ServiceIQS(
                name=svc.title(),
                fotos_avaliadas=avaliadas,
                fotos_nc=nc,
                fotos_conforme=conforme,
                nc_pct=nc / avaliadas,
                conforme_pct=conforme / avaliadas,
            )
        )
    return out


def _ic_rows_from_inspections(df: pd.DataFrame, services: list[str]) -> list[ServiceIC]:
    out: list[ServiceIC] = []
    if df.empty:
        return out
    for svc in services:
        sub = df[df["service"] == svc]
        total = len(sub)
        if total == 0:
            continue
        conf = int(sub["conforme_count"].sum())
        out.append(ServiceIC(name=svc.title(), ic_pct=conf / total, lvs=total))
    return out


def _iqs_overall_from_inspections(df: pd.DataFrame) -> float | None:
    if df.empty:
        return None
    total = int(df["photo_total"].sum())
    if total == 0:
        return None
    return int(df["photo_conforme"].sum()) / total


def _iso(ts: pd.Timestamp | None) -> str:
    return ts.strftime("%Y-%m-%d") if ts is not None else ""
