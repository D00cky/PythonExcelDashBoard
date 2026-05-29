from pathlib import Path

from openpyxl import Workbook

# Each row: (team, tss, fachada_stage, sinalizacao_stage). Stage values are 'C', 'NC' or 'SF'.
_INSPECTION_ROWS = {
    "ÁGUA": [
        ("JOSIAS ALMEIDA FRANCISCO", "TROCAR RAMAL DE ÁGUA PREVENTIVA", "C", "C"),
        ("JOSIAS ALMEIDA FRANCISCO", "VAZAMENTO DE ÁGUA NO PASSEIO", "C", "NC"),
        ("LAIS RAMOS SOBRAL", "TROCAR RAMAL DE ÁGUA PREVENTIVA", "NC", "NC"),
        ("LAIS RAMOS SOBRAL", "TROCAR RAMAL DE ÁGUA PREVENTIVA", "C", "SF"),
        ("FERNANDO PEREIRA ASSIS DE LIMA MARTINS", "TROCAR RAMAL DE ÁGUA PREVENTIVA", "C", "NC"),
    ],
    "ESGOTO": [
        ("LAIS RAMOS SOBRAL", "TAMPONAR LIGAÇÃO DE ESGOTO", "C", "C"),
        ("FERNANDO PEREIRA ASSIS DE LIMA MARTINS", "VAZAMENTO DE ESGOTO", "NC", "NC"),
        ("JOSIAS ALMEIDA FRANCISCO", "TAMPONAR LIGAÇÃO DE ESGOTO", "C", "C"),
    ],
}


def make_minimal_sabesp(
    tmp_path: Path, *, with_periodo: bool = True, with_inspections: bool = False
) -> Path:
    """Build a minimal SABESP-shape xlsx covering the cells we extract.

    Layout mirrors the real ``DADOS - PIMENTAS`` sheet:
      B4   period string
      B26-G29 per-service IQS rows (Água / Esgoto / Cavalete / Reposição)
      G30  overall IQS conforme %

    ``with_periodo=False`` leaves B4 empty so tests can exercise None handling.
    """
    wb = Workbook()
    wb.active.title = "CAPA"

    wb.create_sheet("DADOS - PIMENTAS")
    dados = wb["DADOS - PIMENTAS"]
    if with_periodo:
        dados["B4"] = "Período: 01/03/2026 à 31/03/2026"

    dados["B10"] = "Tipo de Serviço"
    dados["C10"] = "IC (%)"
    dados["E10"] = "Quantidade de LVs"
    dados["B11"] = "Água"
    dados["C11"] = 1.0
    dados["E11"] = 100
    dados["B12"] = "Esgoto"
    dados["C12"] = 0.5
    dados["E12"] = 50
    dados["B13"] = "Reposição"
    dados["C13"] = 0.1
    dados["E13"] = 1

    dados["B25"] = "Tipo de Serviço"
    dados["C25"] = "Fotos Avaliadas"
    dados["D25"] = "Fotos NC"
    dados["E25"] = "Fotos Conforme"
    dados["F25"] = "Não Conforme (%)"
    dados["G25"] = "Conforme (%)"

    _write_iqs_row(dados, 26, "Água", 182, 91, 91, 0.5, 0.5)
    _write_iqs_row(dados, 27, "Esgoto", 33, 16, 17, 0.484848, 0.515151)
    _write_iqs_row(dados, 28, "Cavalete", 73, 13, 60, 0.178082, 0.821917)
    _write_iqs_row(dados, 29, "Reposição", 31, 6, 25, 0.193548, 0.806451)
    dados["B30"] = "Total"
    dados["C30"] = 319
    dados["D30"] = 126
    dados["E30"] = 193
    dados["F30"] = 0.339120
    dados["G30"] = 0.660880

    wb.create_sheet("ÁGUA")
    wb.create_sheet("ESGOTO")

    if with_inspections:
        for sheet_name, rows in _INSPECTION_ROWS.items():
            ws = wb[sheet_name]
            ws["A1"] = "Unidade Executante"
            ws["C1"] = "Descrição TSS"
            ws["N1"] = "EQUIPE"
            ws["O1"] = "FACHADA"
            ws["Q1"] = "SINALIZAÇÃO"
            for i, (team, tss, fachada, sinalizacao) in enumerate(rows, start=2):
                ws[f"A{i}"] = "TEST"
                ws[f"C{i}"] = tss
                ws[f"N{i}"] = team
                ws[f"O{i}"] = fachada
                ws[f"Q{i}"] = sinalizacao

    path = tmp_path / "sabesp.xlsx"
    wb.save(path)
    return path


def _write_iqs_row(ws, row: int, name: str, avaliadas, nc, conforme, nc_pct, conforme_pct):
    ws[f"B{row}"] = name
    ws[f"C{row}"] = avaliadas
    ws[f"D{row}"] = nc
    ws[f"E{row}"] = conforme
    ws[f"F{row}"] = nc_pct
    ws[f"G{row}"] = conforme_pct
