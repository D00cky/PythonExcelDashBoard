import pandas as pd
from openpyxl import Workbook

from app.core.templates.sabesp_pimentas import (
    SabespPimentasTemplate,
    _failing_inspections,
    _os_conformity_summary,
    _stage_code_counts,
    _top_reason,
    _top_tss,
)
from tests.fixtures.sabesp_minimal import make_minimal_sabesp


def test_extract_team_detail_returns_empty_dict_when_team_absent(tmp_path):
    path = make_minimal_sabesp(tmp_path, with_inspections=True)

    detail = SabespPimentasTemplate().extract_team_detail(path, "NOBODY")

    assert detail == {}


def test_extract_team_detail_aggregates_per_service_os_counts(tmp_path):
    path = make_minimal_sabesp(tmp_path, with_inspections=True)

    detail = SabespPimentasTemplate().extract_team_detail(path, "JOSIAS ALMEIDA FRANCISCO")

    assert set(detail) == {"ÁGUA", "ESGOTO"}
    # JOSIAS in ÁGUA: row (C,C) → conforme; row (C,NC) → not conforme.
    agua = detail["ÁGUA"]
    assert agua["inspecoes"] == 2
    assert agua["conforme"] == 1
    assert agua["nao_conforme"] == 1
    assert agua["conforme"] + agua["nao_conforme"] == agua["inspecoes"]
    # With_nc / with_sf are informational subsets of nao_conforme.
    assert agua["with_nc"] == 1
    assert agua["with_sf"] == 0
    # Fixture has no Observação columns → no reasons can be extracted
    assert agua["top_nc_reason"] is None
    assert agua["top_sf_reason"] is None


def test_extract_team_detail_tss_summary_lists_distinct_services(tmp_path):
    path = make_minimal_sabesp(tmp_path, with_inspections=True)

    detail = SabespPimentasTemplate().extract_team_detail(path, "JOSIAS ALMEIDA FRANCISCO")

    assert dict(detail["ÁGUA"]["tss_summary"]) == {
        "TROCAR RAMAL DE ÁGUA PREVENTIVA": 1,
        "VAZAMENTO DE ÁGUA NO PASSEIO": 1,
    }


def test_extract_team_detail_trims_whitespace_in_team_name(tmp_path):
    path = make_minimal_sabesp(tmp_path, with_inspections=True)

    detail = SabespPimentasTemplate().extract_team_detail(path, "  JOSIAS ALMEIDA FRANCISCO  ")

    assert "ÁGUA" in detail


def test_extract_team_detail_top_nc_reason_picks_most_common_observation(tmp_path):
    path = _write_observations_xlsx(
        tmp_path,
        [
            ("ALICE", "SERV1", "NC", "telhado quebrado"),
            ("ALICE", "SERV1", "NC", "telhado quebrado"),
            ("ALICE", "SERV1", "NC", "muro pichado"),
            ("ALICE", "SERV1", "C", None),
            ("ALICE", "SERV1", "SF", "sem foto da fachada"),
        ],
    )

    detail = SabespPimentasTemplate().extract_team_detail(path, "ALICE")

    assert detail["ÁGUA"]["top_nc_reason"] == ("telhado quebrado", 2)
    assert detail["ÁGUA"]["top_sf_reason"] == ("sem foto da fachada", 1)


def test_top_reason_returns_most_common_string():
    assert _top_reason(["a", "b", "a", "c", "a"]) == ("a", 3)
    assert _top_reason([]) is None


def test_top_tss_returns_empty_when_column_missing():
    assert _top_tss(pd.DataFrame({"x": [1]}), None) == []


def test_top_tss_limits_to_top_n():
    sub = pd.DataFrame({"tss": ["a", "a", "b", "b", "c"]})
    assert _top_tss(sub, "tss", n=2) == [("a", 2), ("b", 2)]


def test_stage_code_counts_returns_empty_when_no_stage_cols():
    assert _stage_code_counts(pd.DataFrame(), []) == {}


def test_os_conformity_summary_handles_no_stage_cols():
    assert _os_conformity_summary(pd.DataFrame(), []) == {
        "conforme": 0,
        "nao_conforme": 0,
        "with_nc": 0,
        "with_sf": 0,
    }


def test_os_conformity_summary_counts_at_row_level():
    # 4 rows: (C,C)→conforme; (C,NC)→fail/with_nc; (SF,C)→fail/with_sf;
    # (NC,SF)→fail/with_nc/with_sf.
    sub = pd.DataFrame({"FACHADA": ["C", "C", "SF", "NC"], "SINAL": ["C", "NC", "C", "SF"]})
    assert _os_conformity_summary(sub, ["FACHADA", "SINAL"]) == {
        "conforme": 1,
        "nao_conforme": 3,
        "with_nc": 2,
        "with_sf": 2,
    }


def test_extract_team_detail_failing_inspections_lists_each_bad_os(tmp_path):
    path = _write_observations_xlsx(
        tmp_path,
        [
            ("ALICE", "TROCAR RAMAL", "NC", "telhado quebrado"),
            ("ALICE", "VAZAMENTO", "C", None),
            ("ALICE", "TROCAR RAMAL", "SF", "sem foto da fachada"),
        ],
        os_numbers=[1001, 1002, 1003],
    )

    detail = SabespPimentasTemplate().extract_team_detail(path, "ALICE")
    failing = detail["ÁGUA"]["failing_inspections"]

    assert [f["tss"] for f in failing] == ["TROCAR RAMAL", "TROCAR RAMAL"]
    assert [f["os"] for f in failing] == ["1001", "1003"]
    codes = [s["code"] for f in failing for s in f["failed_stages"]]
    assert codes == ["NC", "SF"]
    observations = [s["observation"] for f in failing for s in f["failed_stages"]]
    assert observations == ["telhado quebrado", "sem foto da fachada"]


def test_failing_inspections_os_field_blank_when_column_missing(tmp_path):
    path = _write_observations_xlsx(
        tmp_path,
        [
            ("ALICE", "T1", "NC", "x"),
        ],
    )

    failing = SabespPimentasTemplate().extract_team_detail(path, "ALICE")["ÁGUA"][
        "failing_inspections"
    ]
    assert failing[0]["os"] == ""


def test_failing_inspections_returns_empty_when_no_stages():
    assert _failing_inspections(pd.DataFrame(), [], None) == []


def _write_observations_xlsx(tmp_path, rows, os_numbers=None):
    wb = Workbook()
    wb.active.title = "CAPA"
    wb.create_sheet("DADOS - PIMENTAS")
    ws = wb.create_sheet("ÁGUA")
    ws["A1"] = "EQUIPE"
    ws["B1"] = "Descrição TSS"
    ws["C1"] = "FACHADA"
    ws["D1"] = "Observação FACHADA"
    if os_numbers is not None:
        ws["E1"] = "Número OS"
    for i, (team, tss, stage, observation) in enumerate(rows, start=2):
        ws[f"A{i}"] = team
        ws[f"B{i}"] = tss
        ws[f"C{i}"] = stage
        ws[f"D{i}"] = observation
        if os_numbers is not None:
            ws[f"E{i}"] = os_numbers[i - 2]
    path = tmp_path / "team_obs.xlsx"
    wb.save(path)
    return path
