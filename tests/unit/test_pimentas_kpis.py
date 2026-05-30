from openpyxl import load_workbook

from app.core.templates.pimentas import PimentasTemplate
from tests.fixtures.pimentas_minimal import make_minimal_pimentas


def test_extract_periodo_strips_label_prefix(tmp_path):
    wb = load_workbook(make_minimal_pimentas(tmp_path))

    result = PimentasTemplate().extract_periodo(wb)

    assert result == "01/03/2026 à 31/03/2026"


def test_extract_periodo_returns_none_when_cell_empty(tmp_path):
    wb = load_workbook(make_minimal_pimentas(tmp_path, with_periodo=False))

    assert PimentasTemplate().extract_periodo(wb) is None


def test_extract_iqs_by_service_returns_one_row_per_known_service(tmp_path):
    wb = load_workbook(make_minimal_pimentas(tmp_path))

    rows = PimentasTemplate().extract_iqs_by_service(wb)

    assert {r.name for r in rows} == {"Água", "Esgoto", "Cavalete", "Reposição"}

    agua = next(r for r in rows if r.name == "Água")
    assert agua.fotos_avaliadas == 182
    assert agua.fotos_nc == 91
    assert agua.fotos_conforme == 91
    assert agua.conforme_pct == 0.5
    assert agua.nc_pct == 0.5


def test_extract_iqs_overall_returns_total_row_conforme_pct(tmp_path):
    wb = load_workbook(make_minimal_pimentas(tmp_path))

    assert PimentasTemplate().extract_iqs_overall(wb) == 0.660880


def test_extract_ic_by_service_returns_three_services(tmp_path):
    wb = load_workbook(make_minimal_pimentas(tmp_path))

    rows = PimentasTemplate().extract_ic_by_service(wb)

    by_name = {r.name: r for r in rows}
    assert set(by_name) == {"Água", "Esgoto", "Reposição"}
    assert by_name["Água"].ic_pct == 1.0
    assert by_name["Água"].lvs == 100
    assert by_name["Esgoto"].ic_pct == 0.5
    assert by_name["Reposição"].lvs == 1
