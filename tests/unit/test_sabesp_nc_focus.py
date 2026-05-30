from datetime import datetime

import pandas as pd
from openpyxl import Workbook

from app.core.templates.sabesp_pimentas import (
    SabespPimentasTemplate,
    top_observations,
)


def _new_workbook_with_capa():
    wb = Workbook()
    wb.active.title = "CAPA"
    wb.create_sheet("DADOS - PIMENTAS")
    return wb


def _write_failures_xlsx(tmp_path):
    """ÁGUA sheet with 3 NC and 2 SF outcomes plus observation columns."""
    wb = Workbook()
    wb.active.title = "CAPA"
    wb.create_sheet("DADOS - PIMENTAS")
    ws = wb.create_sheet("ÁGUA")
    ws["A1"] = "EQUIPE"
    ws["B1"] = "Descrição TSS"
    ws["C1"] = "FACHADA"
    ws["D1"] = "Observação FACHADA"
    ws["E1"] = "SINALIZAÇÃO"
    ws["F1"] = "Observação SINALIZAÇÃO"
    ws["G1"] = "Data Início Execução"
    rows = [
        ("ALICE", "T1", "NC", "telhado quebrado", "C", None, datetime(2026, 3, 5)),
        (
            "ALICE",
            "T1",
            "NC",
            "telhado quebrado",
            "SF",
            "câmera descarregada",
            datetime(2026, 3, 6),
        ),
        ("BOB", "T2", "C", None, "NC", "placa caída", datetime(2026, 3, 7)),
        ("BOB", "T2", "SF", "celular sem espaço", "C", None, datetime(2026, 3, 8)),
        ("ALICE", "T1", "C", None, "C", None, datetime(2026, 3, 9)),
    ]
    for i, (team, tss, fa, fa_obs, sn, sn_obs, dt) in enumerate(rows, start=2):
        ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"], ws[f"D{i}"] = team, tss, fa, fa_obs
        ws[f"E{i}"], ws[f"F{i}"], ws[f"G{i}"] = sn, sn_obs, dt
    path = tmp_path / "failures.xlsx"
    wb.save(path)
    return path


def test_extract_stage_failures_returns_one_row_per_failing_cell(tmp_path):
    path = _write_failures_xlsx(tmp_path)

    failures = SabespPimentasTemplate().extract_stage_failures(path)

    # 3 NC + 2 SF across the two stage columns = 5 failures total.
    assert len(failures) == 5
    assert set(failures["code"]) == {"NC", "SF"}
    assert set(failures["stage"]) == {"FACHADA", "SINALIZAÇÃO"}
    assert set(failures["team"]) == {"ALICE", "BOB"}


def test_extract_stage_failures_pairs_observation_with_correct_stage(tmp_path):
    path = _write_failures_xlsx(tmp_path)

    failures = SabespPimentasTemplate().extract_stage_failures(path)

    fa_nc = failures[(failures["stage"] == "FACHADA") & (failures["code"] == "NC")]
    assert sorted(fa_nc["observation"].tolist()) == ["telhado quebrado", "telhado quebrado"]
    sn_nc = failures[(failures["stage"] == "SINALIZAÇÃO") & (failures["code"] == "NC")]
    assert sn_nc["observation"].tolist() == ["placa caída"]


def test_top_observations_returns_most_common_for_code(tmp_path):
    path = _write_failures_xlsx(tmp_path)
    failures = SabespPimentasTemplate().extract_stage_failures(path)

    top_nc = top_observations(failures, "NC")
    assert top_nc[0] == {
        "observation": "telhado quebrado",
        "stages": "FACHADA",
        "count": 2,
    }

    top_sf = top_observations(failures, "SF")
    # SF observations: 'câmera descarregada' (FACHADA, but cell is SF) +
    # 'celular sem espaço' (paired with FACHADA SF).
    assert {r["observation"] for r in top_sf} == {
        "câmera descarregada",
        "celular sem espaço",
    }


def test_top_observations_skips_empty_and_returns_empty_for_no_failures():
    empty = pd.DataFrame(
        columns=["service", "stage", "code", "observation", "team", "tss", "start_date"]
    )
    assert top_observations(empty, "NC") == []


def test_build_top_failing_stages_handles_empty_df():
    fig = SabespPimentasTemplate().build_top_failing_stages(pd.DataFrame(columns=["stage", "code"]))
    assert fig.layout.annotations[0].text.startswith("Sem etapas")


def test_build_worst_teams_filters_small_samples():
    df = pd.DataFrame(
        {
            "team": ["A"] * 5 + ["B"] * 2,
            "nao_conforme_count": [1, 1, 1, 0, 0, 1, 1],
        }
    )

    fig = SabespPimentasTemplate().build_worst_teams(df, min_inspections=3)

    # Only team A meets the 3-inspection threshold.
    bar = fig.data[0]
    assert list(bar.y) == ["A"]


# ---------------------------------------------------------------------------
# Failure modes: bad / partial / pathological inputs
# ---------------------------------------------------------------------------


def test_extract_stage_failures_returns_empty_for_sheet_without_equipe(tmp_path):
    """Sheet that lacks the EQUIPE column must not produce ghost rows."""
    wb = _new_workbook_with_capa()
    ws = wb.create_sheet("ÁGUA")
    ws["A1"], ws["B1"] = "FACHADA", "Observação FACHADA"
    ws["A2"], ws["B2"] = "NC", "telhado quebrado"
    path = tmp_path / "no_team.xlsx"
    wb.save(path)

    failures = SabespPimentasTemplate().extract_stage_failures(path)

    assert failures.empty
    assert list(failures.columns) == [
        "service",
        "stage",
        "code",
        "observation",
        "team",
        "tss",
        "start_date",
    ]


def test_extract_stage_failures_drops_rows_with_blank_equipe(tmp_path):
    """Rows whose EQUIPE cell is empty are dropped — no anonymous failures."""
    wb = _new_workbook_with_capa()
    ws = wb.create_sheet("ÁGUA")
    ws["A1"], ws["B1"] = "EQUIPE", "FACHADA"
    ws["A2"], ws["B2"] = "ALICE", "NC"
    ws["A3"], ws["B3"] = None, "NC"  # blank team — should be ignored
    path = tmp_path / "blank_team.xlsx"
    wb.save(path)

    failures = SabespPimentasTemplate().extract_stage_failures(path)

    assert failures["team"].tolist() == ["ALICE"]


def test_extract_stage_failures_handles_missing_tss_and_date_columns(tmp_path):
    """Without TSS / date columns the extractor still works, with empty defaults."""
    wb = _new_workbook_with_capa()
    ws = wb.create_sheet("ÁGUA")
    ws["A1"], ws["B1"], ws["C1"] = "EQUIPE", "FACHADA", "Observação FACHADA"
    ws["A2"], ws["B2"], ws["C2"] = "ALICE", "NC", "telhado quebrado"
    path = tmp_path / "no_tss_no_date.xlsx"
    wb.save(path)

    failures = SabespPimentasTemplate().extract_stage_failures(path)

    assert len(failures) == 1
    assert failures.iloc[0]["tss"] == ""
    assert pd.isna(failures.iloc[0]["start_date"])


def test_extract_stage_failures_normalises_whitespace_in_stage_cells(tmp_path):
    """' NC ' must be recognised as NC — Excel users routinely leak spaces."""
    wb = _new_workbook_with_capa()
    ws = wb.create_sheet("ÁGUA")
    ws["A1"], ws["B1"] = "EQUIPE", "FACHADA"
    ws["A2"], ws["B2"] = "ALICE", " NC "
    ws["A3"], ws["B3"] = "BOB", " SF"
    path = tmp_path / "whitespace.xlsx"
    wb.save(path)

    failures = SabespPimentasTemplate().extract_stage_failures(path)

    assert sorted(failures["code"].tolist()) == ["NC", "SF"]


def test_extract_stage_failures_excludes_conforme_and_na_cells(tmp_path):
    """Only NC / SF count as failures — C and NA must not slip through."""
    wb = _new_workbook_with_capa()
    ws = wb.create_sheet("ÁGUA")
    ws["A1"], ws["B1"] = "EQUIPE", "FACHADA"
    for i, code in enumerate(["C", "NA", "NC", "SF"], start=2):
        ws[f"A{i}"], ws[f"B{i}"] = "ALICE", code
    path = tmp_path / "mixed_codes.xlsx"
    wb.save(path)

    failures = SabespPimentasTemplate().extract_stage_failures(path)

    assert sorted(failures["code"].tolist()) == ["NC", "SF"]


def test_extract_stage_failures_stage_without_observation_column_is_kept(tmp_path):
    """Stage cell without a following 'Observação ...' column → observation None."""
    wb = _new_workbook_with_capa()
    ws = wb.create_sheet("ÁGUA")
    # FACHADA has no observation column next to it (next col is another stage).
    ws["A1"], ws["B1"], ws["C1"] = "EQUIPE", "FACHADA", "SINALIZAÇÃO"
    ws["A2"], ws["B2"], ws["C2"] = "ALICE", "NC", "C"
    path = tmp_path / "no_obs_col.xlsx"
    wb.save(path)

    failures = SabespPimentasTemplate().extract_stage_failures(path)

    assert len(failures) == 1
    assert failures.iloc[0]["observation"] is None


def test_extract_stage_failures_blank_observation_cell_is_none_not_nan_string(tmp_path):
    """An NC with no observation must produce observation=None — never the string 'nan'."""
    wb = _new_workbook_with_capa()
    ws = wb.create_sheet("ÁGUA")
    ws["A1"], ws["B1"], ws["C1"] = "EQUIPE", "FACHADA", "Observação FACHADA"
    ws["A2"], ws["B2"], ws["C2"] = "ALICE", "NC", None
    path = tmp_path / "blank_obs.xlsx"
    wb.save(path)

    failures = SabespPimentasTemplate().extract_stage_failures(path)

    assert len(failures) == 1
    obs = failures.iloc[0]["observation"]
    assert obs is None or (isinstance(obs, float) and pd.isna(obs))


def test_top_observations_skips_whitespace_only_observations():
    """A spaces-only observation has zero signal — must not surface as a 'top motivo'."""
    df = pd.DataFrame(
        {
            "code": ["NC", "NC", "NC"],
            "observation": ["   ", "telhado quebrado", "telhado quebrado"],
            "stage": ["FACHADA", "FACHADA", "FACHADA"],
        }
    )

    top = top_observations(df, "NC")

    assert top == [{"observation": "telhado quebrado", "stages": "FACHADA", "count": 2}]


def test_top_observations_does_not_leak_across_codes():
    """NC results must not include observations whose code is SF, and vice versa."""
    df = pd.DataFrame(
        {
            "code": ["NC", "SF"],
            "observation": ["telhado quebrado", "câmera descarregada"],
            "stage": ["FACHADA", "FACHADA"],
        }
    )

    nc = top_observations(df, "NC")
    sf = top_observations(df, "SF")

    assert {r["observation"] for r in nc} == {"telhado quebrado"}
    assert {r["observation"] for r in sf} == {"câmera descarregada"}


def test_top_observations_caps_results_at_n():
    """When more distinct observations exist than n, only the top n are returned."""
    df = pd.DataFrame(
        {
            "code": ["NC"] * 5,
            "observation": [f"motivo {i}" for i in range(5)],
            "stage": ["FACHADA"] * 5,
        }
    )

    top = top_observations(df, "NC", n=3)

    assert len(top) == 3


def test_top_observations_joins_distinct_stages_for_same_text():
    """Same observation across two stages → 'FACHADA, SINALIZAÇÃO' (sorted, distinct)."""
    df = pd.DataFrame(
        {
            "code": ["NC", "NC", "NC"],
            "observation": ["mesma falha"] * 3,
            "stage": ["SINALIZAÇÃO", "FACHADA", "FACHADA"],
        }
    )

    top = top_observations(df, "NC")

    assert top[0]["stages"] == "FACHADA, SINALIZAÇÃO"
    assert top[0]["count"] == 3


def test_top_observations_returns_empty_when_code_has_no_failures():
    """Requesting SF on a DF with only NC rows must return [] (no KeyError)."""
    df = pd.DataFrame(
        {
            "code": ["NC", "NC"],
            "observation": ["a", "b"],
            "stage": ["FACHADA", "FACHADA"],
        }
    )

    assert top_observations(df, "SF") == []


def test_build_top_failing_stages_handles_only_nc_no_sf():
    """If no SF rows exist, the SF trace should still render with zeros — no KeyError."""
    df = pd.DataFrame(
        {
            "stage": ["FACHADA", "FACHADA", "SINALIZAÇÃO"],
            "code": ["NC", "NC", "NC"],
        }
    )

    fig = SabespPimentasTemplate().build_top_failing_stages(df)

    nc_trace = next(t for t in fig.data if t.name == "Não Conforme")
    sf_trace = next(t for t in fig.data if t.name == "Sem Foto")
    assert all(v == 0 for v in sf_trace.x)
    assert sum(nc_trace.x) == 3


def test_build_top_failing_stages_trims_to_top_n_largest():
    """With more stages than top_n, only the top_n by total count are kept."""
    rows = []
    for i, n in enumerate([1, 2, 3, 4, 5]):
        rows.extend({"stage": f"E{i}", "code": "NC"} for _ in range(n))

    fig = SabespPimentasTemplate().build_top_failing_stages(pd.DataFrame(rows), top_n=3)

    nc_trace = next(t for t in fig.data if t.name == "Não Conforme")
    # ascending sort then tail(3) → bottom-to-top order is [E2, E3, E4].
    assert list(nc_trace.y) == ["E2", "E3", "E4"]
    assert list(nc_trace.x) == [3, 4, 5]


def test_build_worst_teams_returns_empty_state_when_all_teams_below_threshold():
    """No team meets min_inspections → specific empty-state message, not a crash."""
    df = pd.DataFrame(
        {
            "team": ["A", "A", "B"],
            "nao_conforme_count": [1, 0, 1],
        }
    )

    fig = SabespPimentasTemplate().build_worst_teams(df, min_inspections=5)

    assert fig.data == ()
    assert "≥ 5 inspeções" in fig.layout.annotations[0].text


def test_build_worst_teams_returns_empty_state_when_no_team_has_failures():
    """Teams meet the threshold but all are 100% conforme → distinct empty-state message."""
    df = pd.DataFrame(
        {
            "team": ["A"] * 4,
            "nao_conforme_count": [0, 0, 0, 0],
        }
    )

    fig = SabespPimentasTemplate().build_worst_teams(df, min_inspections=3)

    assert fig.data == ()
    assert fig.layout.annotations[0].text == "Nenhuma equipe com falhas no período"


def test_build_worst_teams_orders_worst_at_top_of_y_axis():
    """Plotly y-axis grows upward → the worst team must be last in the y list."""
    df = pd.DataFrame(
        {
            "team": ["A"] * 4 + ["B"] * 4 + ["C"] * 4,
            # A: 25% fail, B: 75% fail, C: 50% fail. Worst = B.
            "nao_conforme_count": [1, 0, 0, 0, 1, 1, 1, 0, 1, 1, 0, 0],
        }
    )

    fig = SabespPimentasTemplate().build_worst_teams(df, min_inspections=3)

    bar = fig.data[0]
    assert list(bar.y) == ["A", "C", "B"]
    assert bar.y[-1] == "B"


def test_build_worst_teams_empty_dataframe_uses_no_inspections_message():
    """Distinct empty-state copy for the empty-DF vs threshold-not-met cases."""
    fig = SabespPimentasTemplate().build_worst_teams(
        pd.DataFrame(columns=["team", "nao_conforme_count"])
    )

    assert fig.layout.annotations[0].text == "Sem inspeções registradas"
