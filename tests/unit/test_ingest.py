from openpyxl import Workbook

from app.core.ingest import read_workbook


def test_read_workbook_returns_dict_keyed_by_sheet_name(tmp_path):
    wb = Workbook()
    wb.active.title = "First"
    wb.create_sheet("Second")
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    result = read_workbook(path)

    assert set(result.keys()) == {"First", "Second"}


def test_read_workbook_preserves_portuguese_diacritics(tmp_path):
    wb = Workbook()
    wb.active.title = "ÁGUA"
    wb["ÁGUA"]["A1"] = "Reposição"
    wb["ÁGUA"]["A2"] = "Conformidade — Não Conforme"
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    result = read_workbook(path)

    assert "ÁGUA" in result
    assert "Reposição" in result["ÁGUA"].columns
    assert result["ÁGUA"].iloc[0, 0] == "Conformidade — Não Conforme"
