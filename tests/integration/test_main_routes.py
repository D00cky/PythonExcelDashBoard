import io
from html.parser import HTMLParser
from pathlib import Path

_EXTERNAL_PREFIXES = ("http://", "https://", "//")


def _external_resource_attrs(html: str) -> list[tuple[str, str]]:
    """Return [(tag, url), ...] for every <script src> / <link href> that loads off-host."""

    class Finder(HTMLParser):
        def __init__(self) -> None:
            super().__init__()
            self.refs: list[tuple[str, str]] = []

        def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
            d = dict(attrs)
            if tag == "script":
                src = d.get("src") or ""
                if src.startswith(_EXTERNAL_PREFIXES):
                    self.refs.append((tag, src))
            elif tag == "link":
                href = d.get("href") or ""
                if href.startswith(_EXTERNAL_PREFIXES):
                    self.refs.append((tag, href))

    finder = Finder()
    finder.feed(html)
    return finder.refs


def test_index_route_serves_upload_form(client):
    response = client.get("/")

    assert response.status_code == 200
    assert b"<form" in response.data
    assert b'type="file"' in response.data


def test_upload_rejects_missing_file_with_400(client):
    response = client.post("/upload")

    assert response.status_code == 400


def test_upload_rejects_non_xlsx_extension_with_400(client):
    response = client.post(
        "/upload",
        data={"file": (io.BytesIO(b"not an xlsx"), "notes.txt")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 400


def test_upload_persists_file_and_303_redirects_to_dashboard(client, app):
    payload = b"xlsx-bytes-stand-in"

    response = client.post(
        "/upload",
        data={"file": (io.BytesIO(payload), "report.xlsx")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 303
    assert response.location.startswith("/dashboard/")

    upload_id = response.location.removeprefix("/dashboard/")
    saved = Path(app.instance_path) / "uploads" / f"{upload_id}.xlsx"
    assert saved.read_bytes() == payload


def test_dashboard_renders_sabesp_kpis_and_two_figures(client, tmp_path):
    from tests.fixtures.sabesp_minimal import make_minimal_sabesp

    payload = make_minimal_sabesp(tmp_path, with_inspections=True).read_bytes()
    upload_response = client.post(
        "/upload",
        data={"file": (io.BytesIO(payload), "report.xlsx")},
        content_type="multipart/form-data",
    )

    response = client.get(upload_response.location)

    assert response.status_code == 200
    body = response.data.decode("utf-8")
    # Fixture inspections span 2026-03-05..2026-03-29; that becomes the period
    # (CAPA's stated 'à 31/03/2026' is overridden when inspection dates exist).
    assert "05/03/2026 à 29/03/2026" in body
    assert "66.1%" in body
    assert "Polo Pimentas" in body
    # 5 top-level figures + 4 services × (team + tss) per-service charts = 13
    assert body.count('class="plotly-graph-div"') == 13
    assert "Listas de Verificação" not in body
    assert "Fotos Avaliadas" in body
    assert "Inspeções Avaliadas" in body
    assert "Conformidade por Equipe" in body
    assert "Conformidade por TSS" in body


def test_dashboard_returns_404_for_unknown_id(client):
    response = client.get("/dashboard/" + "0" * 32)

    assert response.status_code == 404


def _upload_minimal(client, tmp_path) -> str:
    from tests.fixtures.sabesp_minimal import make_minimal_sabesp

    payload = make_minimal_sabesp(tmp_path, with_inspections=True).read_bytes()
    upload = client.post(
        "/upload",
        data={"file": (io.BytesIO(payload), "report.xlsx")},
        content_type="multipart/form-data",
    )
    return upload.location.removeprefix("/dashboard/")


def test_download_md_returns_markdown_summary(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(f"/download/{upload_id}?fmt=md")

    assert response.status_code == 200
    assert response.mimetype == "text/markdown"
    assert f"dashboard-{upload_id}.md" in response.headers["Content-Disposition"]
    body = response.data.decode("utf-8")
    assert body.startswith("# Dashboard SABESP")
    assert "Polo Pimentas" in body
    assert "05/03/2026 à 29/03/2026" in body


def test_download_xlsx_returns_openxml_workbook(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(f"/download/{upload_id}?fmt=xlsx")

    assert response.status_code == 200
    assert "spreadsheetml.sheet" in response.headers["Content-Type"]
    assert response.data[:2] == b"PK"


def test_download_pdf_returns_pdf_bytes(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(f"/download/{upload_id}?fmt=pdf")

    assert response.status_code == 200
    assert response.mimetype == "application/pdf"
    assert response.data[:4] == b"%PDF"


def test_download_docx_returns_office_document(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(f"/download/{upload_id}?fmt=docx")

    assert response.status_code == 200
    assert "wordprocessingml.document" in response.headers["Content-Type"]
    assert response.data[:2] == b"PK"


def test_download_returns_400_for_unsupported_format(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(f"/download/{upload_id}?fmt=txt")

    assert response.status_code == 400


def test_team_detail_renders_breakdown_for_known_team(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(
        f"/dashboard/{upload_id}/team?name=JOSIAS+ALMEIDA+FRANCISCO"
    )

    assert response.status_code == 200
    body = response.data.decode("utf-8")
    assert "JOSIAS ALMEIDA FRANCISCO" in body
    assert "Polo Pimentas" in body


def test_team_detail_returns_400_when_name_param_missing(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(f"/dashboard/{upload_id}/team")

    assert response.status_code == 400


def test_team_detail_returns_404_for_unknown_upload_id(client):
    response = client.get("/dashboard/" + "0" * 32 + "/team?name=X")

    assert response.status_code == 404


def test_team_detail_returns_404_when_team_absent_from_sheets(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(f"/dashboard/{upload_id}/team?name=NOBODY")

    assert response.status_code == 404


def test_dashboard_shows_date_filter_form_with_available_range(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(f"/dashboard/{upload_id}")

    body = response.data.decode("utf-8")
    # Fixture inspections span 2026-03-05..2026-03-29
    assert 'id="start"' in body
    assert 'id="end"' in body
    assert "2026-03-05" in body  # available_start in form
    assert "2026-03-29" in body  # available_end in form
    assert "(filtrado)" not in body  # not filtered by default


def test_dashboard_date_filter_restricts_period_and_inspections(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(
        f"/dashboard/{upload_id}?start=2026-03-15&end=2026-03-25"
    )

    body = response.data.decode("utf-8")
    assert response.status_code == 200
    assert "(filtrado)" in body
    # Within 2026-03-15..2026-03-25 the surviving rows span 15/03 (ESGOTO row)
    # to 25/03 (ÁGUA / FERNANDO row).
    assert "15/03/2026 à 25/03/2026" in body
    assert 'value="2026-03-15"' in body
    assert 'value="2026-03-25"' in body


def test_dashboard_warns_when_inspection_span_exceeds_60_days(client, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "CAPA"
    wb.create_sheet("DADOS - PIMENTAS")
    ws = wb.create_sheet("ÁGUA")
    wb.create_sheet("ESGOTO")
    ws["A1"] = "Descrição TSS"
    ws["B1"] = "Data Início Execução"
    ws["C1"] = "EQUIPE"
    ws["D1"] = "FACHADA"
    from datetime import datetime
    ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "T1", datetime(2026, 1, 5), "ALICE", "C"
    ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "T1", datetime(2026, 4, 5), "ALICE", "C"
    path = tmp_path / "wide.xlsx"
    wb.save(path)

    upload = client.post(
        "/upload",
        data={"file": (io.BytesIO(path.read_bytes()), "wide.xlsx")},
        content_type="multipart/form-data",
    )
    upload_id = upload.location.removeprefix("/dashboard/")

    response = client.get(f"/dashboard/{upload_id}")
    body = response.data.decode("utf-8")
    assert "dia/mês trocados" in body
    assert "warning-banner" in body


def test_dashboard_ignores_invalid_date_params(client, tmp_path):
    upload_id = _upload_minimal(client, tmp_path)

    response = client.get(f"/dashboard/{upload_id}?start=not-a-date&end=also-bad")

    assert response.status_code == 200
    body = response.data.decode("utf-8")
    assert "(filtrado)" not in body
